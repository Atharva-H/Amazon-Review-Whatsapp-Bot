from openpyxl.xml.constants import MAX_ROW
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
import keyboard
from urllib.parse import quote
import time
import openpyxl as opyxl
from os import system, name


def clear():

    # for windows
    if name == 'nt':
        _ = system('cls')

    # for mac and linux(here, os.name is 'posix')
    else:
        _ = system('clear')


pro_start = time.time()
wb = opyxl.load_workbook('data.xlsx')
sheet = wb.active
l = sheet.max_row
length = l-1
n_s = 0
#amz_link = "https://www.amazon.in/review/create-review/?ie=UTF8&channel=glance-detail&asin="
# update css selector if you have any issues
css_selector = "#main > footer > div.vR1LG._3wXwX.copyable-area > div._2A8P4 > div > div._2_1wd.copyable-text.selectable-text"

driver = webdriver.Chrome(ChromeDriverManager().install())

# first call without delay in order to scan qr code
driver.get("https://web.whatsapp.com")
keyboard.wait('`')
clear()

# message to be sent to everyone, you can also read it as a dict from a file with ph nos as keys
msg = ''' 
Thank you so much for your purchase of Paricott Paper cups from Amazon! 
A little of your time and a few brief words would go a long way to help other customers make a decision!

Simply click on the link below to give us star rating & post your review.
https://www.amazon.in/review/create-review/?ie=UTF8&channel=glance-detail&asin='''
msg_arr = []
for n in range(n_s, l-1):
    names = sheet.cell(row=2+n, column=2)
    #msgtemp = "Hi "+str(names.value)+", "+msg
    msgtemp = "Hi "+", "+msg
    # url-encode the message, use other functios for handling dictionaries, not recommended
    msgtemp = quote(msgtemp)
    msg_arr.append(msgtemp)
    # print(msg_arr[n])
    # keyboard.wait('`')


url_list = []

for n in range(n_s, l-1):

    num = sheet.cell(row=2+n, column=3)
    asin = sheet.cell(row=2+n, column=8)
    url_list.append("https://web.whatsapp.com/send?phone=91" +
                    str(num.value) + "&text=" + str(msg_arr[n]) + str(asin.value))
    # print(url_list[n])
    # keyboard.wait('`')

count = 1
failed = 0

send_c = 0
send = False

for url in url_list:
    start = time.time()
    driver.get(url)
    time.sleep(3)
    print("~~~~~~~~~~")
    print("Sending to buyer no.=", count)

    for i in range(0, 5):
        try:
            print("try-", i+1)
            driver.find_element_by_css_selector(
                css_selector).send_keys(Keys.RETURN)
            driver.execute_script("window.onbeforeunload = function() {};")
            print("########## Sent")
            send = True
            send_c += 1
            time.sleep(5)
            break
        except:
            print("Except")
            time.sleep(3)

    if send == False:
        failed += 1
        print("@@@@@@@@@@ Unable to send")
    send = False
    count += 1
    # print(url)
    end = time.time()
    print("Time taken =", "{:.3f}".format(end - start))


pro_end = time.time()
print("~~~~~~~~~~")

print("---------------------------------------------------")
print("                       REPORT                      ")
print("Send File run successfully")
print("Total Time taken     =", "{:.2f}".format(
    (pro_end - pro_start)/60), "min")
print("Avg Time taken       =", "{:.2f}".format(
    (pro_end - pro_start)/length), "sec")
print("sent successfully    =", send_c)
print("Total Unable to send =", failed)
print("Total Text Attempted =", length)
print("---------------------------------------------------")
